#!/usr/bin/env python3
"""Shared I/O and parsing helpers for insect-fitness skill scripts."""

from __future__ import annotations

import csv
import hashlib
import json
import math
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable


def file_sha256(path: str | Path) -> str:
    digest = hashlib.sha256()
    with open(path, "rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def normalize_token(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip().casefold()
    return re.sub(r"[\s_\-—–/\\:：;；,.，。()（）\[\]{}]+", "", text)


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, float) and math.isnan(value):
        return True
    return isinstance(value, str) and not value.strip()


def clean_scalar(value: Any) -> Any:
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def trim_rows(rows: list[list[Any]]) -> list[list[Any]]:
    trimmed = [list(row) for row in rows]
    while trimmed and all(is_blank(value) for value in trimmed[-1]):
        trimmed.pop()
    if not trimmed:
        return []
    width = max((len(row) for row in trimmed), default=0)
    while width > 0 and all(width > len(row) or is_blank(row[width - 1]) for row in trimmed):
        width -= 1
    return [row[:width] + [None] * max(0, width - len(row)) for row in trimmed]


def load_tables(path: str | Path) -> tuple[dict[str, list[list[Any]]], dict[str, Any]]:
    source = Path(path).expanduser().resolve()
    suffix = source.suffix.casefold()
    metadata: dict[str, Any] = {
        "source_file": str(source),
        "sha256": file_sha256(source),
        "format": suffix.lstrip("."),
        "merged_ranges": {},
        "formula_counts": {},
    }

    if suffix in {".csv", ".tsv"}:
        delimiter = "\t" if suffix == ".tsv" else ","
        with source.open("r", encoding="utf-8-sig", newline="") as handle:
            rows = [row for row in csv.reader(handle, delimiter=delimiter)]
        name = source.stem[:31] or "Sheet1"
        return {name: trim_rows(rows)}, metadata

    if suffix not in {".xlsx", ".xlsm"}:
        raise ValueError(f"Unsupported input format: {suffix}. Use XLSX, XLSM, CSV, or TSV.")

    try:
        import openpyxl  # type: ignore
    except ImportError as exc:
        raise RuntimeError(
            "XLSX input requires openpyxl. Use the Codex bundled Python runtime or export the sheet as CSV."
        ) from exc

    workbook = openpyxl.load_workbook(source, read_only=False, data_only=False)
    tables: dict[str, list[list[Any]]] = {}
    for worksheet in workbook.worksheets:
        rows = [[cell.value for cell in row] for row in worksheet.iter_rows()]
        tables[worksheet.title] = trim_rows(rows)
        metadata["merged_ranges"][worksheet.title] = [str(item) for item in worksheet.merged_cells.ranges]
        metadata["formula_counts"][worksheet.title] = sum(
            1
            for row in worksheet.iter_rows()
            for cell in row
            if isinstance(cell.value, str) and cell.value.startswith("=")
        )
    workbook.close()
    return tables, metadata


def select_sheet(tables: dict[str, list[list[Any]]], sheet: str | None) -> tuple[str, list[list[Any]]]:
    if sheet:
        if sheet not in tables:
            raise KeyError(f"Worksheet {sheet!r} not found. Available: {', '.join(tables)}")
        return sheet, tables[sheet]
    if len(tables) != 1:
        raise ValueError(f"Input contains {len(tables)} worksheets; specify --sheet or set mapping.sheet.")
    return next(iter(tables.items()))


def make_headers(row: Iterable[Any]) -> list[str]:
    headers: list[str] = []
    seen: dict[str, int] = {}
    for index, value in enumerate(row, start=1):
        base = str(value).strip() if not is_blank(value) else f"unnamed_{index}"
        count = seen.get(base, 0) + 1
        seen[base] = count
        headers.append(base if count == 1 else f"{base}__{count}")
    return headers


def row_is_repeated_header(row: list[Any], headers: list[str]) -> bool:
    values = [normalize_token(value) for value in row[: len(headers)]]
    expected = [normalize_token(value.split("__", 1)[0]) for value in headers]
    nonempty = [(left, right) for left, right in zip(values, expected) if left or right]
    return bool(nonempty) and sum(left == right for left, right in nonempty) / len(nonempty) >= 0.8


def parse_date_value(value: Any, default_year: int | None = None) -> tuple[str | None, str | None]:
    if is_blank(value):
        return None, None
    if isinstance(value, datetime):
        return value.date().isoformat(), None
    if isinstance(value, date):
        return value.isoformat(), None
    if isinstance(value, (int, float)):
        return None, "numeric_date_requires_confirmation"

    text = str(value).strip()
    formats = [
        "%Y-%m-%d",
        "%Y/%m/%d",
        "%Y.%m.%d",
        "%m-%d-%Y",
        "%m/%d/%Y",
        "%m-%d-%y",
        "%m/%d/%y",
    ]
    for fmt in formats:
        try:
            return datetime.strptime(text, fmt).date().isoformat(), None
        except ValueError:
            pass

    chinese = re.fullmatch(r"(?:(\d{4})年)?(\d{1,2})月(\d{1,2})日?", text)
    if chinese:
        year_text, month_text, day_text = chinese.groups()
        year = int(year_text) if year_text else default_year
        if year is None:
            return None, "date_missing_year"
        try:
            return date(year, int(month_text), int(day_text)).isoformat(), None
        except ValueError:
            return None, "invalid_date"
    return None, "unparsed_date"


def parse_number_value(value: Any) -> tuple[float | int | None, str | None]:
    if is_blank(value):
        return None, None
    if isinstance(value, bool):
        return None, "boolean_not_numeric"
    if isinstance(value, int):
        return value, None
    if isinstance(value, float):
        if math.isnan(value):
            return None, None
        return int(value) if value.is_integer() else value, None
    text = str(value).strip()
    if re.fullmatch(r"[-+]?\d+", text):
        return int(text), None
    if re.fullmatch(r"[-+]?(?:\d+\.\d*|\d*\.\d+)", text):
        return float(text), None
    return None, "unparsed_numeric"


def read_csv_records(path: str | Path) -> list[dict[str, str]]:
    with Path(path).open("r", encoding="utf-8-sig", newline="") as handle:
        return list(csv.DictReader(handle))


def write_csv(path: str | Path, fieldnames: list[str], records: Iterable[dict[str, Any]]) -> None:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    with destination.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for record in records:
            writer.writerow({key: clean_scalar(record.get(key)) for key in fieldnames})


def write_json(path: str | Path, payload: Any) -> None:
    destination = Path(path)
    destination.parent.mkdir(parents=True, exist_ok=True)
    destination.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def quantile(values: list[float], probability: float) -> float | None:
    finite = sorted(value for value in values if value is not None and math.isfinite(value))
    if not finite:
        return None
    if len(finite) == 1:
        return finite[0]
    position = (len(finite) - 1) * probability
    lower = math.floor(position)
    upper = math.ceil(position)
    if lower == upper:
        return finite[lower]
    fraction = position - lower
    return finite[lower] * (1 - fraction) + finite[upper] * fraction
